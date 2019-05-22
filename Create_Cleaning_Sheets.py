# -*- coding: utf-8 -*-
"""
Created on Tue Apr  2 10:12:03 2019

@author: ndr15
"""

#!/usr/bin/python3

import urllib3
import sys
import os
import subprocess
import glob
import io
urllib3.disable_warnings()
from redcap import Project, RedcapError
from requests import post
from pathlib import Path
import smtplib
import datetime
import pycurl
import certifi
import json
import argparse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.pagebreak import Break
from copy import copy


#  new comment

parser = argparse.ArgumentParser(description='RedCap REST client')
parser.add_argument('records_of_interest', metavar='ID', type=str, nargs='+',
                    help='a list of subject IDs to fetch metadata from')
parser.add_argument('--all', dest='export_disabled', action='store_const',
                    const=True, default=False,
                    help='export all subject sessions even when marked as disabled (default: exclude disabled)')

args = parser.parse_args()

records_of_interest = args.records_of_interest



# fetch API key from ~/.redcap-key ... don't keep in the source
key_filename = os.path.expanduser('~') + '/.redcap-key'
if not os.path.isfile(key_filename):
    print('redcap key file {} not found'.format(key_filename))
    sys.exit(1)
api_key = open(key_filename, 'r').read().strip()



api_url = 'https://externalredcap.isd.kcl.ac.uk/api/'
project = Project(api_url, api_key)

infile=load_workbook('DataCheckingTemplate.xlsx')
source=infile['Checking Lists']
outfile_name='checking_batch.xlsx'
records_of_interest=[]
for row in source:
    if row[0].value=='Participant Number' :
        if row[1].value !=None:
            outfile_name=row[1].value.strip()+'.xlsx'          # output file name
    
    if row[0].value!=None:
        if row[0].value[:2].strip()=='CC':
            records_of_interest.append(row[0].value.strip())
            

# get the metadata, form to event mapping and data for this record
#records_of_interest=['CC00106XX07']
meta = project.export_metadata()
fem = project.export_fem()
big_data=project.export_records(records=records_of_interest,
                              format='json')

# build the metadata dictionary so we can find our variables easily
meta_dict={} 
i=0
for var in meta:
    meta_dict[var['field_name']]=i
    i +=1
    
# build a list of events that we have for this participant
    
event_list=['enrolment_arm_1','baby_born_arm_1','fetal_scan_arm_1','neonatal_scan_arm_1']
              


# build output file


source=infile['enrolment_arm_1']




# big loop.  Going through once per participant

for participant in records_of_interest:
    # build data as a list of the events for this participant only
    data=[] 
    for event in big_data:
        if event['participationid']==participant:
            data.append(event)
            
    # now we're going to use the template sheets to build output file with
    # all the relevant sections and the right number of them for scans
    # No data at this time but the output created will be used as both an 
    # input and an output in the data population phase        
        
    source=infile['enrolment_arm_1']        # we'll always have enrolment and baby born
    target=infile.copy_worksheet(source) # copy over the enrolemnt+baby+born # so just copy the whole works
    
    target.cell(row=1,column=3,value=participant)  # write the participant number (hard coded row 1, col 3)
    target.title=participant # output sheet
    # preserve page margins
    target.page_margins.left = source.page_margins.left
    target.page_margins.right = source.page_margins.right
    target.page_margins.top = source.page_margins.top
    target.page_margins.bottom = source.page_margins.bottom
    target.print_title_rows='1:8'
    
    
    scans=['fetal_scan_arm_1','neonatal_scan_arm_1']   # we're going to look for both
    
    for event_of_interest in scans:
        source=infile[event_of_interest]
        for event in data:              # now look through data to see how many of each
            
            if event['redcap_event_name']==event_of_interest and event['scan_disabled'] !='1' : #right event
                out_bottom=target.max_row               # where to start writing on the output
                
                # very inelegant code to copy the cells and the formats
                for i in range(1,source.max_row+1):
                    for j in range(1,source.max_column+1):
                        in_cell=source.cell(row=i,column=j)
                        out_cell=target.cell(row=out_bottom+i,column=j,value=in_cell.value)
                        target.cell(row=out_bottom+i,column=j,value=in_cell.value)._style=copy(in_cell._style)
                        
                # now we need to write out the event_number
                
                new_out_bottom = target.max_row
                for i in range(out_bottom+1,new_out_bottom+1):
                    in_cell=target.cell(row=i,column=1)
                    if in_cell.value == event_of_interest:
                        target.cell(row=i,column=3,value=event['redcap_repeat_instance'])
                     
 
    
    
    
    inrow=0
    event={}                            # dummy end event
    event_of_interest=''
    event_no = 0                    # flag number of events identified in Excel

    for row in target:

        inrow +=1                       # increment the counter
        var=row[0].value                # get the variable or event name
    #    print(var)

        if var in event_list:           # have we found an event to tell us what event we're looking at?
            
            # yes - now check to see whether that event exists in the input
    #        print('found event',var)
    
#            if event_no>0:
##                page_break=Break(id=inrow)
#                print(target.page_breaks)
#                target.page_breaks.append(Break(id=inrow)) # inject page break
            
            event_no +=1
#            print(event_no)
            event_of_interest=''        # clear event_of_interest so we can tell if we found it
            for event in data:
                event_num=''
                if event['redcap_event_name'] ==var and not('dna_sample' in event.values()):
                    
                    if var in scans:
#                        print(var,event['redcap_repeat_instance'],row[2].value)
                        event_num = row[2].value  # get the event number
                    
                    
                    if event_num==event['redcap_repeat_instance']:
                        event_of_interest=var       # found it so set up the event_of_interest
                        break                       # event will be left containing the dictionary for this event
    
    # will fall or break out of loop.  Check if we were successful
            
            if len(event_of_interest)==0:        # did we find it?
                row[2].value=var+' Does not Exist for this Participant'
            
    # if it's a scan we need to fill in the concatenated drugs field
                
            if event_of_interest in scans:
                drugv_bases=['tri1_','tri2_','tri3_','tri4_','xbaby_']
                for drug_var_base in drugv_bases:
    #                print(drug_var_base)
    
                    concat_var_name=drug_var_base +'drugs_concat' # this is the variable name
                    concat_var_value=''                         # initialise it to empty string
                    for j in range(1,21):  # now we're going to loop through each of the individual drugs entries
                        var = drug_var_base+'drug'+str(j)          # build the drug name
    #                    print(var)
                        if var in event:                # just in case it's not there
    #                        print(var)
                            var_value=event[var]        # initialise it to in index 
    #                        print(var_value)
                            if var_value !='':          # ignore it if it's blank
                                if len(concat_var_value)>0 :
                                    concat_var_value +='|'   # introduce separator unless this is 1st
                                var_meta=meta[meta_dict[var]]
                                drop_list=var_meta['select_choices_or_calculations'].split('|')
                                for entry in drop_list:
                                    if entry.split(',')[0].strip()==var_value: # find the entry
                                        
                                        var_value=entry.split(',')[1].strip()
    #                                    print(var_value,entry)
                                        concat_var_value +=var_value            # add it ont
    #                                    print(var_value,concat_var_value)
                                        break                                   # done.  stop searching
                                
                        # finished the enumerated drug.  now need to pick up the 
                    var = drug_var_base +'drug_other'
    #                print(var)
                    if var in event:
                        if event[var] !='':
                            if len(concat_var_value)>0:
                                concat_var_value +='|'
                                concat_var_value +=event[var]
    #                print(concat_var_value)
                    event[concat_var_name]=concat_var_value
                    
    
    # It's not an event.  Might be spare line, comment, header
    
    #    print(var)
        elif var in meta_dict and len(event_of_interest)>0:    # if not a variable just ignore it
            var_meta=meta[meta_dict[var]]
            
    
            
            # now the meaty bit.  Get the value in the appropriate format and passwrite into the template
            
            # dropdown or radio buttons.  Just need to map the entry to the value label
            
            if var_meta['field_type'] == 'dropdown' or var_meta['field_type'] == 'radio' : # dropdown.  Need to map the value into text
                drop_list=var_meta['select_choices_or_calculations'].split('|')
    
                var_value=event[var]            # initialise to the var value.  This protects if entry is blank
                for entry in drop_list:
                    if entry.split(',')[0].strip()==event[var]:
                        var_value=entry.split(',')[1].strip()
    #                    print(var,enrolment_data[var],var_value)
                        break           # this is just breaking out of inner loop
    
            # multiple choice buttons.  Need to search the sub variables and produce a concatentated outputr
            
            elif var_meta['field_type'] == 'checkbox':
                var_value=""
                drop_list=var_meta['select_choices_or_calculations'].split('|')
    
                for entry in drop_list:
                    varx=var+'___'+entry.split(',')[0].strip().replace('-','_')
                    if event[varx]=='1':
                        if len(var_value)>0:
                            var_value=var_value+'|'
                        var_value=var_value+entry.split(',')[1].strip()
            else:
                var_value=event[var]
                
    #        print(var,var_value)       
            row[2].value=var_value
    
    
    
    
    
#    CellFill = PatternFill(fill_type='gray125')
    
    #source['C1'].fill=CellFill           
                  
   

# now output under a fresh name

infile.save(outfile_name)

        

