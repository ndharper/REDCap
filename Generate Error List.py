# -*- coding: utf-8 -*-
"""
Created on Tue Apr  2 10:12:03 2019

@author: ndr15
"""

# !/usr/bin/python3
from redcap import Project, RedcapError
from enum import Enum
import re
# from requests import post
# from pathlib import Path
# import smtplib
import datetime
# import pycurl
# import certifi
# import json
import itertools
import argparse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.pagebreak import Break
from copy import copy
import csv
from pythonds.basic import Stack
from BinaryTree import BinaryTree
from operator import itemgetter
import urllib3
import sys
import os
urllib3.disable_warnings()
# import subprocess
# import glob
# import io


def unpack_choices(m_ent):
    """ return dictionary of value: choice """
    result = {}

    choices = m_ent.split('|')      # break into list of choices
    for choice in choices:
        choice2 = choice.split(',', 1)   # now split the key from the value
        key = choice2[0].strip()      # clean off the blanks
        val = choice2[1].strip()
        result[key] = val

    return result


def process_fields(var, entry, dictionary):
    result = ()             # empty tuple

    var_type = dictionary[var]['field_type']
    m_ent = dictionary[var]
    result_str = ''         # accumulate result

    # checkbox fields.  Will return a tuple(bitmap of options,
    # concatenated string of the texts of set options)
    # if the field is blank will return a tuple of (0,'')
    # but this should never happen if called correctly
    if var_type == 'checkbox':
        # code shared with dropdown and radio
        choices = unpack_choices(m_ent['select_choices_or_calculations'])
        result_code = 0
        i = 0
        for key, desc in choices.items():
            i += 1
            var_key = var + '___' + key
            var_key = var_key.replace('-', '_')  # deal with any hyphens
            var_key = var_key.lower()

            if entry[var_key] == '1':              # is option ticked?
                result_code = result_code * 2 + 1    # generate bitmap
                if len(result_str) > 0:
                    result_str = result_str + '|'
                result_str = result_str + desc

        result = result + (result_code, result_str)
        return result

    # dropdown or radio.  returns tuple (value in variable,text designated by that option)
    # if field is empty ('') will return an empty tuple
    elif var_type == 'dropdown' or var_type == 'radio':
        if entry[var] != '':      # ignore blank
            choices = unpack_choices(m_ent['select_choices_or_calculations'])
            result_str = choices[entry[var]]
            result = result + (entry[var], result_str)
        return result

    # yesno.  returns tuple ('0','No')|('1','Yes') if not blank.  returns empty tuple if blank
    elif var_type == 'yesno':
        if entry[var] != '':
            result = result + (entry[var],)
            if result[0] == '1':
                result = result + ('Yes',)
            else:
                result = result + ('No',)
        return result

    # truefalse.  returns tuple ('0','False')|('1','True') if not blank.  returns empty tuple if blank
    elif var_type == 'truefalse':
        if entry[var] != '':
            result = result + (entry[var],)
            if result[0] == '1':
                result = result + ('True',)
            else:
                result = result + ('False',)

        return result

    # calc or notes fields.  Return tuple with just a single entry, value in the fields.
    # blanks may be valid values for these fields
    elif var_type == 'calc' or var_type == 'notes':
        result = result + (entry[var],)
        return result

    # has to be a text field
    else:

        result = result + (entry[var],)
        # now we need to create the second parameter based on the field validation type

        text_type = m_ent['text_validation_type_or_show_slider_number']

        if text_type == '' or entry[var] == '':
            """
            no validation so it's free form text.  Return as single entry
            tuple
            """
            return result + (entry[var],)

        elif text_type == 'integer':          # integer, return integer value
            result = (entry[var], int(entry[var]))
            """ return raw value and the integer equivalent"""
            return result

        elif text_type == 'number':       # float
            result = (entry[var], float(entry[var]))
            """ return raw value and the floating point equivalent"""
            return result

        elif text_type == 'time':
            result = (entry[var], datetime.datetime.
                      strptime(entry[var], '%H:%M').time())
            return result  # return text and time object

        elif text_type == 'date_dmy':
            result = (entry[var], datetime.datetime.
                      strptime(entry[var], '%Y-%m-%d').date())
            return result  # return text string and date object

        elif text_type == 'date_mdy':  # American date
            result = (entry[var], datetime.datetime.
                      strptime(entry[var], '%Y-%m-%d').date())
            return result  # return text string and date object

        elif text_type == 'date_ymd':  # American date
            result = (entry[var], datetime.datetime.
                      strptime(entry[var], '%Y-%m-%d').date())
            return result  # return text string and date object

        elif text_type == 'datetime_dmy':
            result = (entry[var], datetime.datetime.
                      strptime(entry[var], '%Y-%m-%d %H:%M'))
            return result  # return text string and datetime object

        elif text_type == 'datetime_mdy':  # American date
            result = (entry[var], datetime.datetime.
                      strptime(entry[var], '%Y-%m-%d %H:%M'))
            return result  # return text string and datetime object

        elif text_type == 'datetime_dmy':
            result = (entry[var], datetime.datetime.
                      strptime(entry[var], '%Y-%m-%d %H:%M'))
            return result  # return text string and datetime object
        elif text_type == 'datetime_seconds_dmy':
            result = (entry[var], datetime.datetime.
                      strptime(entry[var], '%Y-%m-%d %H:%M:%S'))
            return result  # return text string and datetime object

        elif text_type == 'datetime_seconds_mdy':  # American date
            result = (entry[var], datetime.datetime.
                      strptime(entry[var], '%Y-%m-%d %H:%M:%S'))
            return result  # return text string and datetime object

        elif text_type == 'datetime_seconds_ymd':
            result = (entry[var], datetime.datetime.
                      strptime(entry[var], '%Y-%m-%d %H:%M:%S'))
            return result  # return text string and datetime object

    return  # should never get here.  Return None = error


_token_re = re.compile(r"""
                        # Constants
((?<![\\])['"])(?P<strc>(?:.(?!(?<![\\])\1))*.?)\1 # 'string' or "string"
|(?P<fcons>\d*\.\d*)                               # floating point number
|(?P<icons>\d+)                                    # fixed point number
|(?P<truec>true)                                   # boolean True
|(?P<falsec>false)                                 # boolean False
                        # REDCap variables.
                        # Allow for optional event and instance
|(?P<Rvar>((\[.*?\])(?:\s*)){1,3})                 # from 1 to 3 [xxx]
                    # Operators
|(?P<ops>not            # unitary not function
|and                    # logical and
|or                     # logical or
|!=                     # logical not equal
|<>                     # logical not equal
|=>                     # logical greater than or equal
|>=                     # logical greater than or equal
|<=                     # logical less than or equal
|=<                     # logical less than or equal
|>                      # logical greater than
|<                      # logical less than
|\+                     # arithmetic or unary add
|\-                     # arithmetic or unary minus
|\*                     # arithmetic multiply
|/                      # arithmetic divide
|\^)                    # raise to power
                    # Functions
|(?P<funcs>datediff     # datediff - 5 parameters
|sum                    # sum - followed by list
|abs                    # absolute value
|min                    # minimum - argument is list
|max                    # max - argument is list
|sqrt                   # square root
|mean                   # mean - argument is list
|median                 # median - argument is list
|stdev                  # standard deviation of list
|roundup                # roundup - argument and places
|rounddown              # rounddown
|round                  # round
|if                     # if(cond,exp_true,exp_false)
|in)                    # isin - followed by list
                    # seperators group
|(?P<seps>\,            # separator func or list memb
|\(                     # open bracket
|\))                    # close bracket
                    # anything else - error
|(?P<errs>\S[a-zA-Z0-9_\.]*)
                      """, re.VERBOSE)


class Token(Enum):
    """ enumeration token types """
    CONST = 0         # string constant (was enclosed in quotes in source)
    RCAP_VAR = 1      # REDCap variable
    OPER = 2          # an operator
    FUNCT = 3         # a function
    SEP = 4           # seperator: comma, ( or )
    ERR = 5           # anything else is an error


def tokenise(s):
    ''' return the tokens one by one'''
    for match in _token_re.finditer(s):
        # various types of constant
        if match.group('strc'):
            yield Token.CONST, match.group('strc')
        elif match.group('fcons'):
            yield Token.CONST, float(match.group('fcons'))
        elif match.group('icons'):
            yield Token.CONST, int(match.group('icons'))
        elif match.group('truec'):
            yield Token.CONST, True
        elif match.group('falsec'):
            yield Token.CONST, False
        # REDCap variables
        # will have variable name encased in square brackets
        # optionally preceeded by an event name, optionally succeeded by
        # and instance number, also encased in square brackets
        elif match.group('Rvar'):   # might get one or more arguments
            # force whitespace between adjacent [..]terms and then split on it
            ss = match.group('Rvar').replace('][', '] [').split()
            r = ()  # empty tuple
            for a in ss:  # fill it with the terms, discarding brackets
                r = r + (a[1:-1],)  # create a tuple of arguments
            yield Token.RCAP_VAR, r
        # operators
        elif match.group('ops'):
            yield Token.OPER, match.group('ops')
        # functions
        elif match.group('funcs'):
            yield Token.FUNCT, match.group('funcs')
        # seperators
        elif match.group('seps'):
            yield Token.SEP, match.group('seps')
        # unrecognised.  error
        else:
            print('Parsing error: expected a token '
                  'but found {} at position {}'.format(match.group('errs'),
                                                       match.span(),
                                                       file=sys.stderr))
            yield Token.ERR, match.group('errs')


# parse the branching logic


def parse_branch(var, meta, my_event, big_data):
    rvar_pat = re.compile(r'\[(.*?)\]')
    str_pat = re.compile(r'\'(.*?)\'')
    paren_pat = re.compile(r'\((.*?)\)')
    operators = {
        '=': '=',
        '>=': '>=',
        '>': '>',
        '<': '<',
        '=>': '>=',
        '<=': '<=',
        '=<': '<=',
        '!=': '!=',
        '<>': '!=',
        'and': 'and',
        'or': 'or',
        ')': ')',
        '(': '('}

    out_list = []
    for entry in meta:
        if entry['field_name'] == var:
            break

    if entry['field_name'] != var:
        # this will happen if the var isn't in the meta data list - error
        return

    if entry['branching_logic'] == '':
        return                          # we've found it but it doesn't have any branching logic

    # now we should have alist of individual terms
    terms = entry['branching_logic'].split()
#
    for term in terms:
        # looking for redcap variable
        matches = rvar_pat.findall(term)
        # we've only got a single term so it must be in current entry
        if len(matches) == 1:
            # subsitute baby(1) = baby___1
            a = paren_pat.sub('___\g<1>', matches[0])

            out_list.append(my_event[a.lower()])

        # if there are two [xxx][yyy] then we have both an event and a variable
        elif len(matches) == 2:
            for entry2 in big_data:  # search through all of big data
                if matches[0] == entry2['redcap_event_name']:
                    # subsitute baby(1) = baby___1
                    a = paren_pat.sub('___\g<1>', matches[1])
                    out_list.append(entry2[a.lower()])
                    break
        else:
            # not a redcap variable.  Look for string argument wrapped in quotes

            matches = str_pat.findall(term)  # should find either 0 or 1
            if len(matches) > 0:
                out_list.append(matches[0])  # will strip the wrapping quotes

        # search for operators or paren
            elif term.strip() in operators:
                out_list.append('%$*&' + term.strip())

            else:
                out_list.append(term)

    return out_list     # should tokenise


def getPrec(node):
    """
    returns the precedence of a node.  Nodes representing operators
    will have values consisting of (operator, precedence) tuples
    operands will contain an elemental value.  Operands are given max
    value precedence
    """
    cur_val = node.getValue()   # this will get the current precedence
    if type(cur_val) == tuple:
        # if node is an operator then value will be a tuple of operator,precedence
        cur_prec = cur_val[1]
    else:
        # if it's an operand then it doesn't have a precedece so force to highest
        cur_prec = 999999999
    return cur_prec


def climbTree(node, prec):
    """
    climb tree until we've found the place to insert a new node
    arguments are the current node and the precedence of the new item.
    If new item precedence is even then we will stop when we find a node with
    a precedence that is less than or equal to the precedence of the new item.
    These are left associative operators
    If the new item precedemce is even then we will climb until we find
    a node that is strictly less than new item.  This places right
    associative operators, e.g. exponentiation, in the right place
    """
    cur_prec = getPrec(node)  # precedence of the current node
    # now have to see if it's left or right associative
    if prec % 2:                    # precedence even or odd?
        while cur_prec > prec:      # odd.  left associative.  Find node <=
            node = node.upTree()    # move up
            cur_prec = getPrec(node)

    else:
        while cur_prec >= prec:     # even.  right associative.  Find node <
            node = node.upTree()    # move up
            cur_prec = getPrec(node)

    return node  # will return pointer to node that we want to insert after


# build parse tree from expression.

def buildParseTree(fplist):
    operators = {
        'and': 2,
        'or': 2,
        '=': 6,
        '!=': 6,
        '>': 6,
        '<=': 6,
        '>': 6,
        '<=': 6,
        '*': 10,
        '/': 10,
        '+': 8,
        '-': 8,
        '^': 11,
    }

    stack = Stack()
    left_bracket = ('(', 0)          # tuple.  0 precedence - lowest

    # dummy an initial left bracket tuple node
    TreeRoot = BinaryTree(left_bracket)
    CurrentNode = TreeRoot

    for entry in fplist:        # loop through each symbol
        if type(entry) != str or str(entry)[:4] != '%$*&':  # operand?
            CurrentNode = CurrentNode.insertBelowCross(entry)

        elif entry == '%$*&(':  # opening bracket?
            CurrentNode = CurrentNode.insertBelowCross(left_bracket)
            stack.push(CurrentNode)  # save so we can get back when we find )

        elif entry == '%$*&)':  # closing bracket
            # pop the opening bracket off the stack and delete it
            # will be left pointing to the parent of the open bracket
            CurrentNode = stack.pop()
            CurrentNode = CurrentNode.deleteNode()

        elif entry[4:] in operators:

            prec = operators[entry[4:]]     # get the precedence
            CurrentNode = climbTree(CurrentNode, prec)
            CurrentNode = CurrentNode.insertBelowCross(
                (entry[4:], operators[entry[4:]]))

        else:
            err_str = 'entry' + entry + 'cannot be processed in module buildParseTree'
            sys.exit(err_str)

    CurrentNode = TreeRoot.deleteNode()  # snip off initial orphan ()
    return CurrentNode

# evaluate parse tree.


def evalParseTree(parse_tree):
    operators = {
        'and': 'logicAnd',
        'or': 'logicOr',
        '=': 'logicEq',
        '!=': 'logicNE',
        '>': 'logicGT',
        '>=': 'logicGE',
        '<': 'logicLT',
        '<=': 'logicLE',
        '*': 'arithMul',
        '/': 'arithDiv',
        '+': 'arithAdd',
        '-': 'arithSub',
        '^': 'arithExp'
    }

    if not isinstance(parse_tree, BinaryTree):
        err_str = 'error: evalParseTree has been called with an argument that is not a BinaryTree'
        sys.exit(err_str)

    val = None  # return value
    if parse_tree.isLeaf():
        return parse_tree.getValue()            # leaf nodes should contail a value

    # we ought to have an operator here
    node_value = parse_tree.getValue()          # this will return a tuple
    if type(node_value) != tuple:
        err_str = 'error: evalParseTree has found a node that ought to be an operator that is not a tuple'
        sys.exit(err_str)
    operator = node_value[0]
    if not operator in operators:
        err_str = 'error: evalParseTree has found a node that ought to be an operator that is not a tuple: ' + operator
        sys.exit(err_str)

    op = operators[operator]          # get the opperator routine
    left = left_arg(parse_tree)     # might be None
    right = right_arg(parse_tree)

    val = eval(op + '(left,right)')
    return val


def logicAnd(left, right):
    return left and right


def logicOr(left, right):
    return left or right

# test that two arguments are equal.  Returns either True or False


def logicEq(left, right):
    if left == right:
        return True

    # they don't match but make sure the problem isn't a type mis-match

    elif type(left) == type(right):      # if they're the same type and they don't match
        return False                    # then they definately don't match

    elif str(left) == str(right):       # try converting both to str.  str(str)==str
        return True

    else:
        return False                    # gave it our best and it didn't match

# test for not equal.  Be lazy and just test for equal and negate


def logicNE(left, right):
    return not(logicEq(left, right))

# all the logical compares will need numeric argument.  Function to ttry to
# convert strings.  Returns tuple(left,right)


def fixType(left, right):
    if type(left) == str:       # string?
        if len(left) == 0:
            return None         # warn to caller that we don't have number
        elif '.' in left:         # yes - does it have a decimal?
            left = float(left)  # yes then try to make it float
        else:
            left = int(left)    # no decimal - try to make it int
    if type(right) == str:
        if len(right) == 0:
            return None
        elif '.' in right:
            right = float(right)
        else:
            right = int(right)
    return (left, right)

# is left >= right


def logicGE(left, right):
    a = fixType(left, right)
    if a == None:
        return False
    return a[0] >= a[1]


def logicLE(left, right):
    a = fixType(left, right)
    if a == None:
        return False
    return a[0] <= a[1]


def logicGT(left, right):
    a = fixType(left, right)
    if a == None:
        return False
    a = fixType(left, right)
    return a[0] > a[1]


def logicLT(left, right):
    a = fixType(left, right)
    if a == None:
        return False
    a = fixType(left, right)
    return a[0] < a[1]


def arithMul(left, right):
    return


def arithDiv(left, right):
    return


def arithAdd(left, right):
    return


def arithSub(left, right):
    return


def arithExp(left, right):
    return


def left_arg(parse_tree):
    left = parse_tree.getLeftChild()
    if not left.isLeaf():                   # if it's not a leaf then call evalParseTree recursively
        return evalParseTree(left)

    return left.getValue()                  # if it's a leaf then just reurn it's value


def right_arg(parse_tree):
    right = parse_tree.getRightChild()
    if not right.isLeaf():                   # if it's not a leaf then call evalParseTree recursively
        return evalParseTree(right)

    return right.getValue()


def clean_participant(data_acc):
    if len(data_acc) == 0:    # any data?
        return []          # no: return empty list

    # look to see if it's a void participant

    for r in data_acc:
        if r['redcap_event_name'] == 'administrative_inf_arm_1':
            if r['void_participant'] == '1':
                return []           # void participant so return an empty list

    # now get rig of any pilot or disabled scans
    scans = 0
    data = []
    for r in data_acc:
        if r['redcap_event_name'] in ['neonatal_scan_arm_1', 'fetal_scan_arm_1']:
            if r['scan_disabled'] == '1' or r['scan_pilot'] == '1':
                continue
            else:
                scans += 1
        data.append(r)

    # if we didn't find any scans then return empty list

    if scans == 0:
        return []
    else:
        return data

 # build the metadata dictionary so we can find our variables easily
# meta_dict={}
# i=0
# for var in meta:
#    meta_dict[var['field_name']]=i
#    i +=1
#
# REDCap will return a list of dictionaries one, dictionary for each event.
# the trouble is that each dictionary will have entries for all the variables
# we asked for, whether they exist in that event or not so we'll have to
# do a complicated set of loops.
# OUTERMOST LOOP: read through the dictionary one variable at a time
# MIDDLE LOOP: get the form name for the variable and loop through the
#    form_event_map looking for entries that reference this form.  There may be
#    more than one if the form is mapped into more than one event.
# INNER LOOP: When we find the form name, loop through the returned data and
#   process the value found in that event.  We will have to loop through all
#    data because there may be multiple instances of the event


def process_participant(data, dictionary, fem):
    is_repeat = {
        'dna_sample': ['baby_born_arm_1',
                       '18_month_assessment_arm_1'],
        'post_scan_events': ['post_scan_event_arm_1']
    }
    for var, rec in dictionary.items():
        form = rec['form_name']

        # now we're going to loop through the form_event_table looking
        # at each event to see if it includes this form
        for form_event in fem:  # fem is a list of dictionaries

            # we've found a reference to this variable's form
            if form == form_event['form']:
                event = form_event['unique_event_name']

    # now go find that event in the REDCap data
    # need to check that the event name is right for this variable
    # and also that the form name matches the redcap repeat instrument
    # latter will only matter for repeating forms in non-repeating events
                for entry in data:  # data is also a list of dictionaries
                    # matched the event.  now check for repeat instruments
                    if entry['redcap_event_name'] == event:
                        if form in is_repeat:
                            # list of events in which this repeats
                            elist = is_repeat[form]
                            if entry['redcap_event_name'] in elist:
                                if form != entry['redcap_repeat_instrument']:
                                    continue  # get out of here
                        elif entry['redcap_repeat_instrument'] != '':
                            continue

                        field_value = process_fields(var, entry, meta)
                        if field_value is None:
                            sys.exit('Parser failure')

                        branch_str = parse_branch(var, meta, entry, data)
                        branch = True  # default is that item isn't hidden
                        if branch_str:
                            tree = buildParseTree(branch_str)
                            branch = evalParseTree(tree)

                        if branch:  # these are the records we want
                            if rec['black_list']:

                                black_list = rec['black_list'].split('|')
                                if len(field_value) > 0:
                                    check = field_value[0]
                                else:
                                    check = ''

                                if check in black_list:
                                    out_write.writerow([entry['participationid'], entry['redcap_event_name'],
                                                        entry['redcap_repeat_instance'], var,
                                                        'Missing Value', field_value])

    return


def _participant_(rec):
    """function to return the participantid for this record"""
    return rec['participationid']


def return_data(big_data):
    """
    return a list of records for a single participant.  need to pass an
    iterator of big_data.  big_data doesn't
    need to be sorted in any particular order but records for a given
    participant must be adjacent
    """
    grouped_data = itertools.groupby(big_data, _participant_)
    for key, gen in grouped_data:

        data = list(gen)
        scans = 0
        selector = []  # will use to filter data later
        for rec in data:
            selector.append(True)
            # void participant
            if args.void and rec['redcap_event_name'] ==\
                    'administrative_inf_arm_1' and \
                    rec['void_participant'] == '1':
                print('participant {} is a void record'.
                      format(rec['participationid']))
                data = []  # chuck everything out
                break  # we're done with this participant

            # check the scan records for void or disabled
            if rec['redcap_event_name'] in ['neonatal_scan_arm_1',
                                            'fetal_scan_arm_1']:
                if args.disabled and rec['scan_disabled'] == '1':
                    selector[-1] = False  # de-select this record
                    print('participant {} {} {} marked disabled'.
                          format(rec['participationid'],
                                 rec['redcap_event_name'],
                                 rec['redcap_repeat_instance']))
                elif args.pilot == 0 and rec['scan_pilot'] == '1':
                    selector[-1] = False  # de-select this record
                    print('participant {} {} {} is a pilot scan'.
                          format(rec['participationid'],
                                 rec['redcap_event_name'],
                                 rec['redcap_repeat_instance']))
                else:
                    scans += 1  # found a good scan

        if len(data) <= 0:
            continue  # would have been a void and we've already printed
        elif args.noscan and scans == 0:
            print('participant {} has no usable scans'.
                  format(rec['participationid']))
            #  data = []  # clear
            continue
        else:
            data = list(itertools.compress(data, selector))
            yield data


parser = argparse.ArgumentParser(description='RedCap REST client')
parser.add_argument('--dictionary', type=str,
                    default='Codebook for Error Check.xlsx [Working Copy]',
                    help='''Dictionary file with variables to check
                    If dictionary file is an Excel workbook then specification
                    can be followed by an optional worksheet name enclosed in
                    square brackets ([]).  If no worksheet name is given then
                    the first sheet in the workbook will be used.  If file
                    does not have an extension .xlx, .xlsx, .xlsb or .xlsm
                    then it will be processed as an ordinary comma seperated
                    test file''')
parser.add_argument('--output', type=argparse.FileType('w'),
                    default='dHCPmissing.txt',
                    help='Output file')
parser.add_argument('--pilot', action='count', default=0,
                    help='''--p means include pilot scans in deciding whether
                    to processs a record but ignore any data in it.
                    --pp means process the data from pilot scans too''')
parser.add_argument('--void', action='store_false',
                    help='include void participant records')
parser.add_argument('--reuse', action='store_true',
                    help='''don\'t reload REDCap DB if it\'s already present in
                    memory.  This will safe time during testing''')
parser.add_argument('--disabled', action='store_false',
                    help='include disabled scan records')
parser.add_argument('--noscan', action='store_false',
                    help='include records with no scans at all')
parser.add_argument('--xlimits', action='store_true',
                    help='''override maximum and minimum limits on variable
                    values in the meta data with values from the dictionary''')
parser.add_argument('records_of_interest', metavar='ID', type=str, nargs='*',
                    help='a list of subject IDs to fetch metadata from')
args = parser.parse_args()


# fetch API key from ~/.redcap-key ... don't keep in the source
key_filename = os.path.expanduser('~') + '/.redcap-key'
if not os.path.isfile(key_filename):
    print('redcap key file {} not found'.format(key_filename))
    sys.exit(1)
api_key = open(key_filename, 'r').read().strip()

api_url = 'https://externalredcap.isd.kcl.ac.uk/api/'  # dHCP specific
project = Project(api_url, api_key)


"""
Check and see if the dictionary file is an Excel spreadsheet.  If it isn't,
assume it's a comma seperated file.  Excel file specification will be of the
form filespec.xl* optionally followed by [Worksheet]. [] are used to contain
the worksheet name beacuse those characters won't occur in the full filespec.
If no worksheet is specified then program will use the active sheet in the
workbook.
If file extension isn't .xls* then assume it's a comma seperated file with the
column headers in the first row
"""

fpat = re.compile(r'\s*(?P<file>.*xls[xmb]?)\s*(\[\s*(?P<sheet>.*)\b)')
match = fpat.match(args.dictionary)

'''
Build a dictionary of dictionaries.  Top level uses the variable name
from 1st column as a key with value equal to a dictionary of all the
other variables in the row with each getting a key equal to the column
heading
'''
dictionary = {}
if match:  # if we matched then it's an Excel file
    print('reading Excel File')
    # read the data dictionary
    dict_cols = {}
    infile = load_workbook(match.group('file'))
    if match.group('sheet'):
        source = infile[match.group('sheet')]
    else:
        source = infile.active
    dg = source.iter_rows(min_row=1, values_only=True)
    header = next(dg, None)  # get the header row

    # if we have a field called 'Ignore' then flag which column
    ig_col = header.index('Ignore') if 'Ignore' in header else -999
    # then loop through the data rows.  Will return a tuple each time
    for rec in dg:
        if ig_col >= 0:
            if rec[ig_col] in ['Yes', 1, '1', True]:
                continue
        dic_entry = {}
        for i in range(1, len(header)):
            if rec[i] is None:
                dic_entry[header[i].lower().replace(' ', '_')] = ''
            else:
                dic_entry[header[i].lower().replace(' ', '_')] = rec[i]

        dictionary[rec[0]] = dic_entry

    infile.close()

else:
    print('reading .csv file')
    with open(args.dictionary, 'r') as infile:
        inreader = csv.reader(infile)
        headers = next(inreader, None)
        ig_col = header.index('Ignore') if 'Ignore' in header else -999
        dictionary = {}
        for rec in inreader:
            if ig_col > 0:
                if rec[ig_col] in ['Yes', 1, '1', True]:
                    continue
            dic_entry = {}
            for i in range(1, len(header)):
                dic_entry[header[i].lower().replace(' ', '_')] = rec[i]
            dictionary[rec[0]] = dic_entry

fields_of_interest = list(dictionary.keys())

"""
get the data from REDCap.  Will retrieve three lists:
    big_data is a list of all the data
    meta is the meta data
    fem is the form event mapping
this takes some time so running interactively we can save time by not reading
again if the data already exists.  Will assume that if big_data exists then
so do meta and fem.  Note that if records_of_interest were specified on the
load and --reuse is specified then the record list doesn't get updated
"""

try:
    big_data  # does it exist?
except NameError:
    args.reuse = False  # force a load
if not args.reuse:
    print('loading REDCap data')
    # fetch API key from ~/.redcap-key ... don't keep in the source
    key_filename = os.path.expanduser('~') + '/.redcap-key'
    if not os.path.isfile(key_filename):
        print('redcap key file {} not found'.format(key_filename),
              file=sys.stderr)
        sys.exit(1)
    api_key = open(key_filename, 'r').read().strip()

    api_url = 'https://externalredcap.isd.kcl.ac.uk/api/'  # dhcp specific
    project = Project(api_url, api_key)

    fields_of_interest = list(dictionary.keys())
    try:
        big_data = project.export_records(fields=fields_of_interest,
                                          records=args.records_of_interest,
                                          format='json')
    except RedcapError:
        print('Redcap export too large', file=sys.stderr)
        sys.exit(1)

    meta = project.export_metadata()                # metadata
    fem = project.export_fem()                      # form event mappings
    """
    copy any meta fields that aren't in the dictionary into it.  If field
    already exists then over-ride it with the exception of
    text_validation_min and max. Those will be spared if --xlimits flag
    is set to allow us to use the limits from the external dictionary.
    This step allows us to use dictionary rather than meta everywhere.
    dictionary is a dictionary of disctionaries rather than a list
    of dictionaries so field lookup is faster

    If the input dictionary was downloaded REDCap codebook then it will
    have the same items as meta but with text descriptions rather than
    variable names.  In fact the only thing we want from the codebook
    is the varaible name plus the added fields, e.g. Ignore and
    black_list and possible min and max value.  When we rtead the header
    we forced to lower case and replaced spaces with '_' so most of the
    fields will be over-written.  A couple won't be but they can just
    sit as garbage in the dictionary as they're never referenced
    """
for row in meta:
    if row['field_name'] in dictionary:
        for key, value in row.items():
            if (key != 'field_name' and
                    not (args.xlimits and
                         key.startswith('text_validation'))):
                dictionary[row['field_name']][key] = value

"""
Check date format.  We can distinguish betweem date and datetime but we don't
have any good way of distinguishing between day-month-year and month-day-year
To address scan the dictionary looking for the text validation values
We have to assume that we won't have a mixture of US and normal dates in the
same project
"""
_American_Date_ = False
for r in dictionary.values():
    if r['field_type'] == 'text':
        ft = r['text_validation_type_or_show_slider_number']
        if ft:
            if ft.find('mdy') >= 0:
                _American_Date_ = True

# open the Excel workbook that has the variables to check

# infile = load_workbook('Codebook for Error Check.xlsx')
# source = infile['Working Copy']

codebook = []

# codebook is deprecated and will be removed
# for row in source.iter_rows(min_row=2, values_only=True):
#     if row[19] != 'Yes':
#         codebook.append(row)

out = open('dHCPmissing.txt', 'w', newline='')
out_write = csv.writer(out, quotechar="'", delimiter='\t')
out_write.writerow(['participant', 'event', 'event_repeat',
                    'variable', 'error', 'value'])


# ibd = iter(big_data)  # create iterable
for data in return_data(big_data):
    if len(data) > 0:                     # have we got any?
        # yes: process them
        process_participant(data, dictionary, fem)

out.close()  # close the output file
