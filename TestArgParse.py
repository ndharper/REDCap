# -*- coding: utf-8 -*-
"""
Created on Wed Sep 11 11:17:05 2019

Test program to play with argparse

@author: ndr15
"""
import urllib3
import sys
import os
import argparse
from redcap import Project, RedcapError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.pagebreak import Break
import csv
import re
import datetime
"""
global flag to indicate that wer have dates in the format month-day-year
as opposed to day-month-year.  We can't tell which format is expected when
we have a literal so we'll parse all the metadat and see if we have any
occurences of mdy.  The assumption is that no project will mix dmy and mdy.
ymd can be mixed with either
"""
AMERICAN_DATE = False
"""
Project specific code.  When we read the database with PyCAP we can identify
repeating events but we can't tell whether an instrument repeats within an
event.  This is a problem because PyCap will return repeating instruments as
seperate instances of the event in which they occur.  These events will still
contain all the fields that we've asked for in the whole project.  These event
records will have redcap_repeat_instrument set to the name of the field and
redcap_repeat_instance will be non-zero.  We will also get an event record for
the part of the event that doesn't come from repeatung instruments.  To make
it even more complicated, an instrument could repeat in one event bit not in
another.  To get round this we have a hard coded dictionary that has entries
for each instrument that does repeat and values that are lists of the vents in
which they do repeat.
"""

IS_REPEAT = {'dna_sample': ['baby_born_arm_1', '18_month_assessment_arm_1'],
             'post_scan_events': ['post_scan_event_arm_1']
             }

"""
Created on Thu Sep 19 13:29:37 2019
@author: nicholas harper
Tokenise a REDCap expression.  These are found in redcap field calculations
and branching logic and are used with external database testing.  Algorithm
derived from Gareth Reese's post
https://codereview.stackexchange.com/questions/186024/basic-equation-tokenizer
looks for additional syntax not found in REDCap:
    strings can be enslosed in single or double quotes and can contain the
    opposite quite unescaped
    includes unary not operator
    includes in function.  Returns true if value is found in list
    allows for operator aliases, e.g <= and =>
"""
from enum import Enum
import re
import sys


_token_re = re.compile(r"""
                        # Constants
((?<![\\])['"])(?P<strc>(?:.(?!(?<![\\])\1))*.?)\1 # 'string' or "string"
|(?P<fcons>\d*\.\d*)                               # floating point number
|(?P<icons>\d+)                                    # fixed point number
|(?P<truec>true)                                   # boolean True
|(?P<falsec>false)                                 # boolean False
                        # REDCap variables.
                        # Allow for optional event and instance
|(?P<Rvar>((\[.*?\])(?:\s*)){1,3})                 # from 1 to 3 [xxx]
                    # Operators
|(?P<ops>not            # unitary not function
|and                    # logical and
|or                     # logical or
|!=                     # logical not equal
|<>                     # logical not equal
|=>                     # logical greater than or equal
|>=                     # logical greater than or equal
|<=                     # logical less than or equal
|=<                     # logical less than or equal
|>                      # logical greater than
|<                      # logical less than
|=                      # logical equal to
|\+                     # arithmetic or unary add
|\-                     # arithmetic or unary minus
|\*                     # arithmetic multiply
|/                      # arithmetic divide
|\^)                    # raise to power
                    # Functions
|(?P<funcs>datediff     # datediff - 5 parameters
|sum                    # sum - followed by list
|abs                    # absolute value
|min                    # minimum - argument is list
|max                    # max - argument is list
|sqrt                   # square root
|mean                   # mean - argument is list
|median                 # median - argument is list
|stdev                  # standard deviation of list
|roundup                # roundup - argument and places
|rounddown              # rounddown
|round                  # round
|if                     # if(cond,exp_true,exp_false)
|in)                    # isin - followed by list
                    # seperators group
|(?P<seps>\,            # separator func or list memb
|\(                     # open bracket
|\))                    # close bracket
                    # anything else - error
|(?P<errs>\S[a-zA-Z0-9_\.]*)
                      """, re.VERBOSE)


class Token(Enum):
    """ enumeration token types """
    CONST = 0         # string constant (was enclosed in quotes in source)
    RCAP_VAR = 1      # REDCap variable
    OPER = 2          # an operator
    FUNCT = 3         # a function
    SEP = 4           # seperator: comma, ( or )
    ERR = 5           # anything else is an error


def tokenise(s):
    ''' return the tokens one by one'''
    for match in _token_re.finditer(s):
        # various types of constant
        if match.group('strc'):
            yield Token.CONST, match.group('strc')
        elif match.group('fcons'):
            yield Token.CONST, float(match.group('fcons'))
        elif match.group('icons'):
            yield Token.CONST, int(match.group('icons'))
        elif match.group('truec'):
            yield Token.CONST, True
        elif match.group('falsec'):
            yield Token.CONST, False
        # REDCap variables
        # will have variable name encased in square brackets
        # optionally preceeded by an event name, optionally succeeded by
        # and instance number, also encased in square brackets
        elif match.group('Rvar'):   # might get one or more arguments
            # force whitespace between adjacent [..]terms and then split on it
            ss = match.group('Rvar').replace('][', '] [').split()
            r = ()  # empty tuple
            for a in ss:  # fill it with the terms, discarding brackets
                r = r + (a[1:-1],)  # create a tuple of arguments
            yield Token.RCAP_VAR, r
        # operators
        elif match.group('ops'):
            yield Token.OPER, match.group('ops')
        # functions
        elif match.group('funcs'):
            yield Token.FUNCT, match.group('funcs')
        # seperators
        elif match.group('seps'):
            yield Token.SEP, match.group('seps')
        # unrecognised.  error
        else:
            print('    : expected a token '
                  'but found {} at position {}'.format(match.group('errs'),
                                                       match.span(),
                                                       file=sys.stderr))
            yield Token.ERR, match.group('errs')


def unpack_choices(m_ent):
    """
    function to unpack the choices of multiple choice filrds: radio, checkbox
    or dropdown.  returns a dictionary with each key being a choice value and
    the corresponding value equal to the text
    """
    result = {}
    choices = m_ent.split('|')      # break into list of choices
    for choice in choices:
        choice2 = choice.split(',', 1)   # now split the key from the value
        key = choice2[0].strip()      # clean off the blanks
        val = choice2[1].strip()
        result[key] = val
    return result


def return_rcap(tup, entry, data, dictionary):
    """
    REDCap variables in branching and tests can be qualified with an event
    and an instance.  Need to point to the right event record before decoding
        1 element - namesd the variable to be retrieved from current record
        2 elements - names an event and the variable
        3 elements - event name, then variable then instance
    """

    if len(tup) == 1:
        return process_fields(tup[0], entry, dictionary)

    elif len(tup) == 2:
        instance = ""
    else:
        instance = tup[2]

    for event in data:
        if event['redcap_event_name'] == tup[0] and \
                            event['redcap_repeat_instance'] == instance:
            break
    return process_fields(tup[1], event, dictionary)


def process_fields(var, entry, dictionary):
    """
    function to process each field from the dictionary and reurn its value
    Different field types require different processing.  Will return a tuple
    with the f=raw value of the field and its interpretation
    """
    result = ()             # empty tuple
    result_str = ''         # accumulate result

    # checkbox variables come from branching as var(item)
    if var.find('(') < 0:  # only checkbox will contain bracket
        var_type = dictionary[var]  # variable type from dictionary
    else:
        var_type = 'checkbox'

    """ process the different field types """
    if var_type == 'checkbox':
        """
        Checkbox fields are complicated because the variable name in meta
        refers to the whole set but the returned data records will contain a
        set of values qualified by the choice.
        """

        if var.find('(') >= 0:  # test again for passed from test
            rpat = re.compile(r'\s*(\S*)\((\S*)\)')
            extract = rpat.search(var)
            choices = unpack_choices(dictionary[extract.group(1)]
                                     ['select_choices_or_calculations'])
            var_key = extract.group(1) + '__' + extract.group(2)
            var_key = var_key.replace('-', '_')  # deal with any hyphens
            result_code = entry[var_key]
            result_str = choices[extract.group(2)]

        else:
            choices = unpack_choices(dictionary[var]
                                     ['select_choices_or_calculations'])
            result_code = 0
            for key, desc in choices.items():
                result_code *= 2  # shift the result bitmap left
                var_key = var + '___' + key
                var_key = var_key.replace('-', '_')  # deal with any hyphens
                var_key = var_key.lower()

                if entry[var_key] == '1':              # is option ticked?
                    result_code += 1                   # set the bit
                    if len(result_str) > 0:
                        result_str = result_str + '|'
                    result_str = result_str + desc

        result = result + (result_code, result_str)
        """
        returned tuple will consist of a result code and a string
        concatenation of all the selected options.  The result code is a
        bitmap where the lowest numbered option is the msb and the last
        option is the lsb.  If no options are ticked then will return (0, '')
        """
        return result

    elif var_type == 'dropdown' or var_type == 'radio':
        """
        dropdown or radio field
        """
        if entry[var] != '':      # ignore blank
            choices = unpack_choices(dictionary[var]
                                               ['select_choices_'
                                                'or_calculations'])
            result_str = choices[entry[var]]
            result = result + (entry[var], result_str)
        """
        returned tuple is the actual value of the variable and it's string
        description.  If the variable is empty will return empty tuple
        """
        return result

    elif var_type == 'yesno':
        if entry[var] != '':
            result = result + (entry[var],)
            if result[0] == '1':
                result = result + ('Yes',)
            else:
                result = result + ('No',)
        """ returns ('0','No') if no, ('1','Yes') if yes
        returns empty tuple if blank
        return result
        """

    elif var_type == 'truefalse':
        if entry[var] != '':
            result = result + (entry[var],)
            if result[0] == '1':
                result = result + ('True',)
            else:
                result = result + ('False',)
        """ returns ('0','False') if flase, ('1','True') if true
        returns empty tuple if blank
        return result
        """
        return result

    elif var_type == 'calc' or var_type == 'notes':
        """
        calc or notes fields.  Return tuple with just a single entry,
        value in the fields. blanks may be valid values for these fields
        """
        result = result + (entry[var],)
        return result

    elif var_type == 'text':
        """
        text can be treated differently depending on the validation type
        """
        text_type = ''
        for m_ent in meta:
            if var == m_ent['field_name']:
                text_type = m_ent['text_validation_type_or_show_slider_number']
                break

        if text_type == '' or entry[var] == '':
            """
            no validation so it's free form text.  Return as single entry
            tuple
            """
            return result + (entry[var],)

        elif text_type == 'integer':          # integer, return integer value
            result = (entry[var], int(entry[var]))
            """ return raw value and the integer equivalent"""
            return result

        elif text_type == 'number':       # float
            result = (entry[var], float(entry[var]))
            """ return raw value and the floating point equivalent"""
            return result

        elif text_type == 'time':
            result = (entry[var], datetime.datetime.
                      strptime(entry[var], '%H:%M').time())
            return result  # return text and time object

        elif text_type == 'date_dmy':
            result = (entry[var], datetime.datetime.
                      strptime(entry[var], '%d-%m-%Y').date())
            return result  # return text string and date object

        elif text_type == 'date_mdy':  # American date
            result = (entry[var], datetime.datetime.
                      strptime(entry[var], '%m-%d-%Y').date())
            return result  # return text string and date object

        elif text_type == 'date_ymd':  # American date
            result = (entry[var], datetime.datetime.
                      strptime(entry[var], '%Y-%m-%d').date())
            return result  # return text string and date object

        elif text_type == 'datetime_dmy':
            result = (entry[var], datetime.datetime.
                      strptime(entry[var], '%d-%m-%Y %H:%M'))
            return result  # return text string and datetime object

        elif text_type == 'datetime_mdy':  # American date
            result = (entry[var], datetime.datetime.
                      strptime(entry[var], '%m-%d-%Y %H:%M'))
            return result  # return text string and datetime object

        elif text_type == 'datetime_dmy':
            result = (entry[var], datetime.datetime.
                      strptime(entry[var], '%Y-%m-%d %H:%M'))
            return result  # return text string and datetime object
        elif text_type == 'datetime_seconds_dmy':
            result = (entry[var], datetime.datetime.
                      strptime(entry[var], '%d-%m-%Y %H:%M:%S'))
            return result  # return text string and datetime object

        elif text_type == 'datetime_seconds_mdy':  # American date
            result = (entry[var], datetime.datetime.
                      strptime(entry[var], '%m-%d-%Y %H:%M:%S'))
            return result  # return text string and datetime object

        elif text_type == 'datetime_seconds_dmy':
            result = (entry[var], datetime.datetime.
                      strptime(entry[var], '%Y-%m-%d %H:%M:%S'))
            return result  # return text string and datetime object

    return  # should ver get here.  Return None = error


def clean_participant(data_acc):
    """
    Garbage cleaning of reords for this participant.  Will get rid of void
    participants, disabled scans and, depending on flag, pilot scans.
    """
    if len(data_acc) == 0:  # any data?
        return []           # no: return empty list

    # look to see if it's a void participant
    if args.void:  # don't do this check if --VOID flag on command
        for r in data_acc:
            if r['redcap_event_name'] == 'administrative_inf_arm_1':
                if r['void_participant'] == '1':
                    return []           # return empty list

    # now get rig of any disabled scans
    if args.disabled:
        for r in data_acc:
            if r['redcap_event_name'] in ['neonatal_scan_arm_1',
                                          'fetal_scan_arm_1']:
                if r['scan_disabled'] == '1':
                    data_acc.remove(r)      # get rid of it

    # now the pilot scans
    if args.pilot == 0:  # default level . throw them out
        for r in data_acc:
            if r['redcap_event_name'] in ['neonatal_scan_arm_1',
                                          'fetal_scan_arm_1']:
                if r['scan_pilot'] == '1':
                    data_acc.remove(r)      # get rid of it

    # now count how many scans we have
    scans = 0
    for r in data_acc:
        if r['redcap_event_name'] in ['neonatal_scan_arm_1',
                                      'fetal_scan_arm_1']:
            scans += 1

    # and throw if we don't have any scans
    if args.noscan:
        if scans == 0:
            return []
    # return the cleaned accumulator
    return data_acc


def process_participant(data, fem, dictionary):
    for var in dictionary:
        form = dictionary[var]['Form Name']      # this is the form name

        # now we're going to loop through the form_event_table looking at
        # each event to see if it includes this form
        for form_event in fem:  # fem is a list of dictionaries

            if form == form_event['form']:
                event = form_event['unique_event_name']

                """
                now go find that event in the REDCap data
                need to check that the event name is right for this variable.
                """
                for entry in data:  # data is also a list of dictionaries
                    if entry['redcap_event_name'] == event:
                        """
                        found the right event.  Now check to make sure it's
                        the right record for repeating/non-repreating
                        instruments
                        """
                        if form in IS_REPEAT:  # this form sometimes repeats
                            elist = IS_REPEAT[form]
                            if entry['redcap_event_name'] in elist:
                                if form != entry['redcap_repeat_instrument']:
                                    continue
                        elif entry['redcap_repeat_instrument'] != '':
                            continue

                        field_value = process_fields(var, entry, dictionary)
                        if field_value is None:  # should return a tuple
                            print('Parsing error in {} variable {}'.
                                  format(entry['participationid'], var),
                                  file=sys.stderr)

                        """
                        now we can do the actual error processing
                        """
                        
                        


#                        branch_str = parse_branch(var, meta, entry,data)
#                        branch = True  # default is that item isn't hidden
#                        if branch_str:
#                            tree = list(tokenise(branch_str))
#                            branch = evalParseTree(tree)
#
#                        if branch: # these are the records in which we are interested
#                            if var_code[18]:
#
#                                black_list=var_code[18].split('|')
#                                if len(field_value)>0:
#                                    check = field_value[0]
#                                else:
#                                    check = ''
#
#                                if check in black_list:
#                                    out_write.writerow([entry['participationid'],entry['redcap_event_name'],\
#                                                        entry['redcap_repeat_instance'],var,\
#                                                        'Missing Value',field_value])


    return


def build_data(ibd):
    """
    generator to return a list of records for the next participant.
    expects to be passed an iterator created from the REDCap data output.
    This doesn't necessarily have to be sorted but all the records for a
    particular participant have to occur together
    """
    data = []  # initialise data buffer
    for rec in ibd:  # loop through the iterator

        # get rid of any disabled scans
        if args.disabled:

            if rec['redcap_event_name'] in ['neonatal_scan_arm_1',
                                            'fetal_scan_arm_1']:
                if rec['scan_disabled'] == '1':
                    continue  # just ignore it

        # now the pilot scans
        if args.pilot == 0:  # default level . throw them out
            if rec['redcap_event_name'] in ['neonatal_scan_arm_1',
                                            'fetal_scan_arm_1']:
                if rec['scan_pilot'] == '1':
                    continue

        if len(data) > 0:  # do we have any data already
            if rec['participationid'] == data[0]['participationid']:
                data.append(rec)  # yes; if we're on same participant, add it
            else:
                # getiing ready to return the data.  Need to check if it's
                data = valid_participant(data)
                if len(data) > 0:
                    yield data
                data = [rec]  # initialise the data buffer for next time
        else:
            data = [rec]  # no previous data so start off with this record
    print(len(data))
    data = valid_participant(data)
    if len(data) > 0:
        yield data  # last one after all the records have been read
   
def valid_participant(data):
    """check if the accumulated data is any good"""
    scans = 0
    void = False
    for r in data:
        if r['redcap_event_name'] in ['neonatal_scan_arm_1',
                                      'fetal_scan_arm_1']:
            scans += 1
        elif r['redcap_event_name'] == 'administrative_inf_arm_1':
            if r['void_participant'] == '1':
                void = True
        if not((args.noscan and scans == 0) or (args.void and void)):
            return data
        else:
            return []

parser = argparse.ArgumentParser(description='RedCap REST client')
parser.add_argument('--dictionary', type=str,
                    default='Codebook for Error Check.xlsx [Working Copy]',
                    help='''Dictionary file with variables to check
                    If dictionary file is an Excel workbook then specification
                    can be followed by an optional worksheet name enclosed in
                    square brackets ([]).  If no worksheet name is given then
                    the first sheet in the workbook will be used.  If file
                    does not have an extension .xlx, .xlsx, .xlsb or .xlsm
                    then it will be processed as an ordinary comma seperated
                    test file''')
parser.add_argument('--output', type=argparse.FileType('w'),
                    default='dHCPmissing.txt',
                    help='Output file')
parser.add_argument('--pilot', action='count', default=0,
                    help='''--p means include pilot scans in deciding whether
                    to processs a record but ignore any data in it.
                    --pp means process the data from pilot scans too''')
parser.add_argument('--void', action='store_false',
                    help='include void participant records')
parser.add_argument('--reuse', action='store_true',
                    help='don\'t reload REDCap DB - saves time when testing')
parser.add_argument('--disabled', action='store_false',
                    help='include disabled scan records')
parser.add_argument('--noscan', action='store_false',
                    help='include records with no scans at all')
parser.add_argument('--xlimits', action='store_true',
                    help='''override maximum and minimum limits on variable
                    values in the meta data with values from the dictionary''')
parser.add_argument('records_of_interest', metavar='ID', type=str, nargs='*',
                    help='a list of subject IDs to fetch metadata from')
args = parser.parse_args()


"""
Check and see if the dictionary file is an Excel spreadsheet.  If it isn't,
assume it's a comma seperated file.  Excel file specification will be of the
form filespec.xl* optionally followed by [Worksheet]. [] are used to contain
the worksheet name beacuse those characters won't occur in the full filespec.
If no worksheet is specified then program will use the active sheet in the
workbook.
If file extension isn't .xls* then assume it's a comma seperated file with the
column headers in the first row
"""

fpat = re.compile(r'\s*(?P<file>.*xls[xmb]?)\s*(\[\s*(?P<sheet>.*)\b)')
match = fpat.match(args.dictionary)

# build the dictionary
dictionary = {}
if match:  # if we matched then it's an Excel file
    print('reading Excel File')
    # read the data dictionary
    dict_cols = {}
    infile = load_workbook(match.group('file'))
    if match.group('sheet'):
        source = infile[match.group('sheet')]
    else:
        source = infile.active
    dg = source.iter_rows(min_row=1, values_only=True)
    header = next(dg, None)
    '''
    Build a dictionary of dictionaries.  Top level uses the variable name
    from ist column as a kdictey with value equal to a dictionary of all the
    other variables in the row with each getting a key equal to the column
    heading
    '''
    ig_col = header.index('Ignore') if 'Ignore' in header else -999

    for rec in dg:
        if ig_col >= 0:
            if rec[ig_col] in ['Yes', 1, True]:
                continue
        dic_entry = {}
        for i in range(1, len(header)):
            dic_entry[header[i]] = rec[i]

        dictionary[rec[0]] = dic_entry

    infile.close()

else:
    print('reading .csv file')
    with open(args.dictionary, 'r') as infile:
        inreader = csv.reader(infile)
        headers = next(inreader, None)
        ig_col = header.index('Ignore') if 'Ignore' in header else -999
        dictionary = {}
        for rec in inreader:
            if ig_col > 0:
                if rec[ig_col] in ['Yes', 1, True]:
                    continue
            dic_entry = {}
            for i in range(1, len(header)):
                dic_entry[header[i]] = rec[i]
            dictionary[rec[0]] = dic_entry

"""
Check date format.  We can distinguish betweem date and datetime but we don't
have any good way of distinguishing between day-month-year and month-day-year
To address scan the dictionary looking for the text validation values
We have to assume that we won't have a mixture of US and normal dates in the
same project
"""

for r in dictionary.values():
    if r['Field Type'] == 'text':
        ft = r['Text Validation Type OR Show Slider Number']
        if ft:
            if ft.find('mdy') >= 0:
                AMERICAN_DATE = True

"""
get the data from REDCap.  Will retrieve three lists:
    big_data is a list of all the data
    meta is the meta data
    fem is the form event mapping
this takes some time so running interactively we can save time by not reading
again if the data already exists.  Will assume that if big_data exists then
so do meta and fem.  Note that if records_of_interest were specified on the
load and --reuse is specified then the record list doesn't get updated
"""

try:
    big_data  # does it exist?
except NameError:
    print('loading REDCap data')
    args.reuse = False  # force a load
if not args.reuse:
    # fetch API key from ~/.redcap-key ... don't keep in the source
    key_filename = os.path.expanduser('~') + '/.redcap-key'
    if not os.path.isfile(key_filename):
        print('redcap key file {} not found'.format(key_filename),
              file=sys.stderr)
        sys.exit(1)
    api_key = open(key_filename, 'r').read().strip()

    api_url = 'https://externalredcap.isd.kcl.ac.uk/api/'  # dhcp specific
    project = Project(api_url, api_key)

    fields_of_interest = list(dictionary.keys())
    try:
        big_data = project.export_records(fields=fields_of_interest,
                                          records=args.records_of_interest,
                                          format='json')
    except RedcapError:
        print('Redcap export too large', file=sys.stderr)
        sys.exit(1)

    meta = project.export_metadata()                # metadata
    fem = project.export_fem()                      # form event mappings
    """
    copy any meta fields that aren't in the dictionary into it.  If field
    already     exists then over-rtide it with teh exception of
    text_validation_min and max.     Those will be spared if --xlimits flag
    is set to allow us to use the limits from the external dictionary.
    This step allows us to use dictionary rather than meta everywhere.
    dictionary is a dictionary of disctionaries rather than a list
    of dictionaries so field lookup is faster
    """
    for row in meta:
        for key, value in row:
            if (key != 'field_name' and
                    not (args.xlimits and key[0:15] == 'text_validation')):
                dictionary[row['field_name'][key]] = value



"""
Main body of the program. For each partipant in turn we need to group all the
event records for that participants.  As part of that stage we will identify
items that don't need to be checked, e.g. only pilot scans, void participant.
"""
currentid = ''      # this the id that we're working on right now
data = []            # build a list of records


# open the output file and write the header
out = open(args.output.name, 'w', newline='')
out_write = csv.writer(out, quotechar="'", delimiter='\t')
out_write.writerow(['participant', 'event', 'event_repeat', 'variable',
                    'error', 'value'])

for record in big_data:
    if record['participationid'] != currentid:   # new participant?
        # Yes -     process the collection we've accumulated
        data = clean_participant(data)        # go do the garbage collection
        if len(data) > 0:                     # have we got any?
            process_participant(data, fem, dictionary)

        currentid = record['participationid']     # new participant
        data = [record]                      # start the list
    else:
        data.append(record)                    # add the record onto the list

data = clean_participant(data)  # process the last participant


if len(data) > 0:                     # have we got any?
    process_participant(data, meta, fem, dictionary)  # yes: process them

out.close()  # close the output file
