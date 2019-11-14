# -*- coding: utf-8 -*-
"""
Created on Tue Apr  2 10:12:03 2019

@author: ndr15
"""

# !/usr/bin/python3
from redcap import Project, RedcapError
from enum import Enum
import re

import datetime
import itertools
import argparse
from openpyxl import Workbook, load_workbook
from copy import copy
import csv
from pythonds.basic import Stack
from BinaryTree import BinaryTree
from operator import itemgetter
import urllib3
import sys
import os
import math
import statistics
urllib3.disable_warnings()


class REDCap_variable:
    """
    object to contain a redcap variable.  Has three attriibutes:
        literal is whatever is actually stored in REDCap.  Will always
           be string even if empty.  If field is a checkbox and the field
           is accessed using the unqualified variable name then lieteral
           will be a text string of '0's or '1's comprising a bitmap of
           all options.  It's little endian so 1st option is leftmost
        value is the content cast into a numeric, boolean or datetime
           object based on the validation rule.
        label is the string representing the contents and will normally
            only be set for checkbox, radio and dropdown
    """

    def __init__(self, lit, value=None, label=None):
        self.literal = lit
        if value:
            self.value = value
        else:
            try:
                self.value = int(lit)  #  see if we have an integer
            except ValueError:
                pass  # we'll try a float
            try:
                self.value = float(lit)
            except ValueError:
                self.value = lit
        self.label = label

    def __str__(self):
        return 'literal: {}; value type: {};\
                value: {}; label: {}'.format(self.literal, type(self.value),
                                             self.value, self.label)


def return_redcap_var(var, entry, dictionary):
    """
    Returns a REDCap_variable object corresponding to the variables data
    Expects to be passed the variable name, the entry from the database
    that has corresponds to the event for this participant containing
    the variable and then the dictionary for the project
    """
    def unpack_choices(m_ent):
        """
        for checkbox, radio and dropdown fields return a list of choices
        """
        result = {}
        choices = m_ent.split('|')      # break into list of choices
        for choice in choices:
            choice2 = choice.split(',', 1)  # now split the key from the value
            key = choice2[0].strip()      # clean off the blanks
            val = choice2[1].strip()
            result[key] = val

        return result

    if var not in dictionary:  # should only happen for checkbox
        """ need to process checkbox variables that can come in in
        either variabale(opt) or variable__option form"""
        # _checkbox_ is global.  Got decalered and compiled in __main__

        match = _checkbox_.search(var)
        if match.group(1) in dictionary:  # bracket form
            m_ent = dictionary[match.group(1)]
            var_key = (match.group(1) + '___'
                       + match.group(2)[0:1].replace('-', '_')
                       + match.group(2)[2:]).lower()
            choice = match.group(2)
        elif match.group(3) in dictionary:  # __form
            m_ent = dictionary[match.group(3)]
            var_key = var
            choice = match.group(4).replace('____', '-___')
        else:
            print('can\'t resolve call to return_redcap_var with'
                  'variable {}'.format(var), file=sys.stderr)
            return None  # signal failure with None return
        result = (entry[var_key], )  # '0' or '1'
        choices = unpack_choices(m_ent['select_choices_or_calculations'])
        field_value = entry[var_key]
        return REDCap_variable(lit=entry[var_key], label=choices[choice])
    var_type = dictionary[var]['field_type']
    m_ent = dictionary[var]  # meta entry

    result_str = ''         # accumulate result

    # checkbox fields.  Will return a bitmap of options,
    # concatenated string of the texts of set options)
    if var_type == 'checkbox':
        # code shared with dropdown and radio
        choices = unpack_choices(m_ent['select_choices_or_calculations'])
        result_code = []  # build bitmap
        i = 0
        for key, desc in choices.items():
            i += 1
            var_key = var + '___' + key
            var_key = var_key.replace('-', '_')  # deal with any hyphens
            var_key = var_key.lower()

            result_code.append(entry[var_key])
            if entry[var_key] == '1':              # is option ticked?
                if len(result_str) > 0:
                    result_str = result_str + '|'
                result_str = result_str + desc

        return REDCap_variable(lit=''.join(result_code), label=result_str)

    # dropdown or radio.  returns tuple (value in variable,text
    # designated by that option).  If field is empty ('') will
    # return an empty tuple
    elif var_type == 'dropdown' or var_type == 'radio':
        if entry[var] != '':      # ignore blank
            choices = unpack_choices(m_ent['select_choices_or_calculations'])
            return REDCap_variable(entry[var], label=choices[entry[var]])
        else:
            return REDCap_variable(entry[var], label='')  # empty

    # yesno
    elif var_type == 'yesno':
        if entry[var] == '':
            return REDCap_variable(entry[var], label='')
        elif entry[var] == '1':
            return REDCap_variable(entry[var], label='Yes')
        else:
            return REDCap_variable(entry[var], label='No')

    # truefalse
    elif var_type == 'truefalse':
        if entry[var] == '':
            return REDCap_variable(entry[var], label='')
        elif entry[var] == '1':
            return REDCap_variable(entry[var], value=True, label='True')
        elif entry[var] == '0':
            return REDCap_variable(entry[var], value=False, label='False')

    # calc or notes fields.  Return tuple with just a single
    # entry, value in the fields. Blanks may be valid values for these fields
    elif var_type == 'calc' or var_type == 'notes':
        return REDCap_variable(entry[var])

    # has to be a text field.  Return the value and the text_type so
    # we can decode it later
    else:
        # now we need to create the second parameter based on the
        # field validation type

        text_type = m_ent['text_validation_type_or_show_slider_number']

        if text_type == '' or entry[var] == '':
            """
            no validation so it's free form text.  Return as single entry
            tuple
            """
            return REDCap_variable(entry[var])

        elif text_type == 'integer':          # integer, return integer value
            return REDCap_variable(entry[var], value=int(entry[var]))

        elif text_type == 'number':       # float
            return REDCap_variable(entry[var], value=float(entry[var]))

        elif text_type == 'time':
            return REDCap_variable(entry[var], value=datetime.datetime.
                                   strptime(entry[var], '%H:%M').time())

        elif text_type in ['date_dmy', 'date_mdy', 'date_ymd']:
            return REDCap_variable(entry[var], datetime.datetime.
                                   strptime(entry[var], '%Y-%m-%d').date())

        elif text_type in ['datetime_mdy', 'datetime_dmy', 'datetime_ymd']:
            return REDCap_variable(entry[var], datetime.datetime.
                                   strptime(entry[var], '%Y-%m-%d %H:%M'))

        elif text_type in ['datetime_seconds_mdy', 'datetime_seconds_dmy',
                           'datetime_seconds_ymd']:
            return REDCap_variable(entry[var], datetime.datetime.
                                   strptime(entry[var], '%Y-%m-%d %H:%M:%S'))

    return  # should never get here.  Return None = error


def decodeRedcapVar(v, entry, dictionary):
    """
    return a REDCap_variable object corresponding to a redcap variable
    specified in branching or test logic
    THIS CODE WILL NEED TO BE UPDATED TO SUPPORT SMART VARIABLES

    """
    if type(v) != tuple:
        print('decodeRedcapVar called with invalid arguement',
              v, file=sys.stderr)
        return
    # need to find how many arguments we have
    if len(v) < 1:
        print('decodeRedcapVar called with invalid arguement',
              v, file=sys.stderr)
        return
    # SELF as a variables means the variable we are testing.
    # included to allow test logic to be used for more than
    # one variable
    if v[0] == 'SELF':
        return return_redcap_var(Current_Variable)
    if len(v) == 1:  # if we only have one term, must be the variable
        return return_redcap_var(v[0], entry, dictionary)

    # we have a specification with an event and maybe an instance
    event = v[0]
    variable = v[1]
    if len(v) >= 3:
        repeat_instance = v[2]
    else:
        repeat_instance = ''

    for e in data:
        if event == e['redcap_event_name']:
            if repeat_instance == e['redcap_repeat_instance']:
                return return_redcap_var(variable, e, dictionary)

    return  # will return None if we get here meaning variable not found


_token_re = re.compile(r"""
                        # Constants
((?<![\\])['"])(?P<strc>(?:.(?!(?<![\\])\1))*.?)\1 # 'string' or "string"
|(?P<fcons>\d*\.\d*)                               # floating point number
|(?P<icons>\d+)                                    # fixed point number
|(?P<truec>true)                                   # boolean True
|(?P<falsec>false)                                 # boolean False
                        # REDCap variables.
                        # Allow for optional event and instance
|(?P<Rvar>((\[.*?\])(?:\s*)){1,3})                 # from 1 to 3 [xxx]
                    # Operators
|(?P<ops>not            # unitary not function.  use !
|and                    # logical and
|or                     # logical or
|!=                     # logical not equal
|<>                     # logical not equal
|=>                     # logical greater than or equal
|>=                     # logical greater than or equal
|<=                     # logical less than or equal
|=<                     # logical less than or equal
|>                      # logical greater than
|<                      # logical less than
|=                      # logical equal to
|!                      # unary logical not
|\*\*                   # raise to power
|\+                     # arithmetic or unary add
|\-                     # arithmetic or unary minus
|\*                     # arithmetic multiply
|/                      # arithmetic divide
|\|\|                   # logical or
|&&)                    # logical and

                    # Functions
|((?P<funcs>datediff    # datediff - 5 parameters
|sum                    # sum - followed by list
|abs                    # absolute value
|min                    # minimum - argument is list
|max                    # max - argument is list
|sqrt                   # square root
|mean                   # mean - argument is list
|median                 # median - argument is list
|stdev                  # standard deviation of list
|roundup                # roundup - argument and places
|rounddown              # rounddown
|round                  # round
|if                     # if(cond,exp_true,exp_false)
|in)                    # isin - followed by list
\s*\()                  # eat the opening bracket after function
                    # seperators group
|(?P<seps>\,            # comma: separator func or list memb
|\(                     # open bracket
|\))                    # close bracket
                        # anything else - error
|(?P<errs>\S[a-zA-Z0-9_\.]*)
                      """, re.VERBOSE)

# regex to parse the conetents of a quoted string

_token_str_ = re.compile(r"""
    (?P<fcons>[+-]?\d*\.\d*)    # floating point number
    |(?P<icons>[+-]?\d+)        # fixed point number
    |(?P<truec>true)            # boolean True
    |(?P<falsec>false)          # boolean False
    |(?P<date3>\s*\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})
    |(?P<date2>\s*\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2})
    |(?P<date1>\s*\d{4}-\d{2}-\d{2})
    |(?P<time>>\s*\d{2}:\d{2})
    """, re.VERBOSE)


class Token(Enum):
    """ enumeration token types """
    CONST = 0         # string constant (was enclosed in quotes in source)
    RCAP_VAR = 1      # REDCap variable
    OPER = 2          # an operator
    FUNCT = 3         # a function
    SEP = 4           # seperator: comma, ( or )
    ERR = 5           # anything else is an error


def tokenise(s):
    ''' return the tokens one by one'''
    for match in _token_re.finditer(s):
        # various types of constant
        if match.group('strc'):  # quoted string
            # try and convert string if we can
            string_cons = _token_str_.search(match.group('strc'))
            if string_cons.group('fcons'):  # float including leadinf +/-
                yield Token.CONST, float(string_cons.group('fcons'))
            elif string_cons.group('icons'):  # integer including leading +/-
                yield Token.CONST, int(string_cons.group('icons'))
            elif string_cons.group('truec'):   # boolean true
                yield Token.CONST, True
            elif string_cons.group('falsec'):  # boolean false
                yield Token.CONST, False
            elif string_cons.group('date3'):  # datetime with HMS
                yield Token.CONST, datetime.datetime.\
                    strptime(string_cons.group('date3', '%Y-%m-%d %H:%M:%S'))

            elif string_cons.group('date2'):  # datetime with HM
                yield Token.CONST, datetime.datetime.\
                    strptime(string_cons.group('date3', '%Y-%m-%d %H:%M'))

            elif string_cons.group('date1'):  # datetime with no time
                yield Token.CONST, datetime.datetime.\
                    strptime(string_cons.group('date3', '%Y-%m-%d'))

            elif string_cons.group('time'):  # datetime with time only
                yield Token.CONST, datetime.datetime.\
                    strptime(string_cons.group('date3', '%H:%M'))

            else:  # can't do anything with it.  Just return as string
                yield Token.CONST, match.group('strc')

        elif match.group('fcons'):
            yield Token.CONST, float(match.group('fcons'))
        elif match.group('icons'):
            yield Token.CONST, int(match.group('icons'))
        elif match.group('truec'):
            yield Token.CONST, True
        elif match.group('falsec'):
            yield Token.CONST, False
        # REDCap variables
        # will have variable name encased in square brackets
        # optionally preceeded by an event name, optionally succeeded by
        # and instance number, also encased in square brackets
        elif match.group('Rvar'):   # might get one or more arguments
            # force whitespace between adjacent [..]terms and then split on it
            ss = match.group('Rvar').replace('][', '] [').split()
            r = ()  # empty tuple
            for a in ss:  # fill it with the terms, discarding brackets
                r = r + (a[1:-1],)  # create a tuple of arguments
            yield Token.RCAP_VAR, r
        # operators
        elif match.group('ops'):
            yield Token.OPER, match.group('ops')
        # functions
        elif match.group('funcs'):
            yield Token.FUNCT, match.group('funcs')
        # seperators
        elif match.group('seps'):
            yield Token.SEP, match.group('seps')
        # unrecognised.  error
        else:
            print('Parsing error: expected a token '
                  'but found {} at position {}'.format(match.group('errs'),
                                                       match.span()),
                  file=sys.stderr)
            yield Token.ERR, match.group('errs')





def getPrec(node):
    """
    returns the precedence of a node.  Nodes representing operators
    will have values consisting of (operator, precedence) tuples
    operands will contain an elemental value.  Operands are given max
    value precedence
    """
    cur_val = node.getValue()   # this will get the current precedence
    if type(cur_val) == tuple:
        # if node is an operator then value will be a tuple of operator,precedence
        cur_prec = cur_val[1]
    else:
        # if it's an operand then it doesn't have a precedece so force to highest
        cur_prec = 999999999
    return cur_prec


# build parse tree from expression.


operators2 = {
    'or': (2, '||'),       # logical or
    '||': (2, '||'),       # logical or
    '&&': (4, '&&'),       # logical and
    'and': (4, '&&'),      # logical and - Alias
    '=': (6, '='),         # logical is equal to
    '!=': (6, '!='),       # logical is not equal to
    '<>': (6, '!='),       # logical is not equal to - Alias
    '>': (6, '>'),         # logical greater than
    '>=': (6, '>='),       # logical greater or equal to
    '=>': (6, '>='),       # logical greater or equal to - Alias
    '<': (6, '<'),         # logical less than
    '<=': (6, '<='),       # logical less than or equal to
    '=<': (6, '<='),       # logical less than or equal to - Alias
    '+': (8, '+'),         # arithmatic addition
    '-': (8, '-'),         # arithmatic subtraction
    '*': (10, '*'),        # arithmatic multiplication
    '/': (10, '/'),        # arithmatic division
    '**': (11, '**'),      # arithmatic exponiation
    '!': (12, '!'),        # logical not
    'not': (12, '!'),      # logical not - Alias
    # unary operators won't be found in input stream but parser will replace
    # binary + or 1 with them where they are found after an opning '(', after
    # another operator or at start of line
    'n': (12, 'n'),        # unary negative
    'p': (12, 'p')          # unary positive - does nothing
}


def parseExpression(s):
    """
    parse the conditional or calculation string by tokenising and
    building a binary tree.  The tree operations are in the module
    BinaryTree
    Returns the tree or None if error
    """
    stack = Stack()
    TreeRoot = BinaryTree('(')  # start tree with dummy "("
    CurrentNode = TreeRoot
    prev_term = None  # will use to distinguish unary and binary + or -
    for term in tokenise(s):
        if term[0] == Token.CONST or term[0] == Token.RCAP_VAR:
            new_node = BinaryTree(term, 999)
            CurrentNode = CurrentNode.addToTree(new_node)
        elif term[0] == Token.FUNCT:
            new_node = BinaryTree(term, 100)
            CurrentNode = CurrentNode.addToTree(new_node)
            CurrentNode.precedence = 0
            stack.push(CurrentNode)
        elif term[0] == Token.OPER:
            if term[1] in '+, -':  # these can be unary or binary
                # will be unary if 1st on a line or 1st after OPER (
                if (not prev_term) or (prev_term[0] == Token.OPER) or\
                        (prev_term[0] == Token.SEP and prev_term[1] == '('):
                    if term[1] == '-':
                        term = (Token.OPER, 'n')  # use the unary
                    elif term[1] == '+':
                        term = (Token.OPER, 'p')  # use the unary
            # replace alias and add precedence
            op_term = (term[0], operators2[term[1]][1])
            prec = operators2[term[1]][0]
            new_node = BinaryTree(op_term, prec)
            CurrentNode = CurrentNode.addToTree(new_node)
        elif term[0] == Token.SEP:
            if term[1] == '(':
                new_node = BinaryTree('(')  # 0 precedence
                CurrentNode = CurrentNode.insertBelowCross(new_node)
                stack.push(CurrentNode)
            elif term[1] == ',':
                CurrentNode = stack.peek()
                CurrentNode = CurrentNode.appendFunc()
            elif term[1] == ')':
                CurrentNode = stack.pop()
                if CurrentNode.getValue()[0] == Token.FUNCT:
                    CurrentNode.precedence = 100
                    CurrentNode = CurrentNode.appendFunc()
                else:
                    CurrentNode = CurrentNode.deleteNode()
        elif term[0] == Token.ERR:
            return  # should already have produced error
        prev_term = term  # so we can check for unary ops after op
    return TreeRoot.deleteNode()  # snip the initial fake '('


def evaluateTree(parse_tree, entry, dictionary):
    print(parse_tree.printTree())
    """
    take the parse tree passed as an argument and evaluate
    it recursively
    """
    operators = {
        '&&': 'logicAnd',  #
        '||': 'logicOr',
        '=': 'logicEq',
        '!=': 'logicNE',
        '!': 'logicNOT',
        '>': 'logicGT',
        '>=': 'logicGE',
        '<': 'logicLT',
        '<=': 'logicLE',
        '*': 'arithMul',
        '/': 'arithDiv',
        '+': 'arithAdd',
        '-': 'arithSub',
        '^': 'arithExp',
        'n': 'unary_negative',
        'p': 'unary_positive'

    }

    functions = {
        'datediff': '_datediff_',
        'sum': '_sum_',
        'abs': '_abs_',
        'min': '_min_',
        'max': '_max_',
        'sqrt': '_sqrt_',
        'mean': '_mean_',
        'median': '_median_',
        'stdev': '_stdev_',
        'roundup': '_roundup_',
        'rounddown': '_rounddown_',
        'round': '_round_',
        'if': '_if_',
        'in': '_in_'
    }

    if not isinstance(parse_tree, BinaryTree):
        return  # return none.  will normally be because we've
        # reached the end of a branch so the child is None

    node_value = parse_tree.getValue()  # should get a token.
    if node_value[0] == Token.CONST:   # constant?
        if not parse_tree.isLeaf():
            print('evaluteTree found a constant node with branches',
                  file=sys.stderr)
        return node_value[1]  # return the value
    elif node_value[0] == Token.RCAP_VAR:
        if not parse_tree.isLeaf():
            print('evaluteTree found a REDCap variable node with branches',
                  file=sys.stderr)
        REDCap_obj = decodeRedcapVar(node_value[1], entry, dictionary)
        return REDCap_obj.value
    elif node_value[0] == Token.OPER:
        # operators take one or two arguments so we'll evaluate
        # left and right branches below this node.  Left branch will
        # be none for a a single argument operator
        left = evaluateTree(parse_tree.getLeftChild(), entry, dictionary)
        right = evaluateTree(parse_tree.getRightChild(), entry, dictionary)
        operator = node_value[1]
        op = operators[operator]          # get the operator routine
#        print(left, operator, right)

        val = eval(op + '(left,right)')
        return val

    elif node_value[0] == Token.FUNCT:
        """
        functions are like operators except that they can have a
        variable number of arguments.  The arguments are contained
        as a tuple of nodes in the left branch.  Some branches may
        not need to be evaluated for all functions, e.g. one of the
        conditions in an if function so we'll defer evaluating until
        the function itself
        """
        function = node_value[1]
        func = functions[function]
        f_args = parse_tree.getLeftChild()  # tuple of trees
        val = eval(func + 'f_args)')
        return val


# parse tree functiuons


def logicAnd(left, right):
    return bool(left) and bool(right)


def logicOr(left, right):
    return bool(left) or bool(right)


def logicNOT(left, right):
    return bool(right)


def logicEq(left, right):
    if left == right:
        return True

    # they don't match but make sure the problem isn't a type mis-match

    elif type(left) == type(right):  # if they're the same type and they don't
        return False                 # match then they definately don't match

    elif str(left) == str(right):  # try converting both to str.  str(str)==str
        return True

    else:
        return False                # gave it our best and it didn't match


# test for not equal.  Be lazy and just test for equal and negate


def logicNE(left, right):
    return not(logicEq(left, right))


def logicGE(left, right):
    return left >= right


def logicLE(left, right):
    return left <= right


def logicGT(left, right):
    return left > right


def logicLT(left, right):
    return left < right


def arithMul(left, right):
    return left * right


def arithDiv(left, right):
    return left / right


def arithAdd(left, right):
    return left + right


def arithSub(left, right):
    return left - right


def arithExp(left, right):
    return left ** right


def unaray_negative(left, right):
    return -right


def unary_positive(left, right):
    return right


"""
functions.  A difference between functions and operators is that an operator
expects the arguments to be fully resolved before the operator is called.
Functions are passed a tuple of branches and the function is responsible
for processing the branches before evaluationg the function.
"""


def _list_args_(args):
    """
    evaluate all the branches returning as a list of values.  Used by the
    sum, mean, median and stdev functions
    """
    arg_list = []
    for a in args:
        arg_list.append(evaluateTree(a))
    return arg_list


# strptime formats
strptime_fmt = {
    'ymd': '%Y-%m-%d',
    'mdy': '%m-%d-%Y',
    'dmy': '%d-%m-%Y'
}

time_fmt = ['', ' %H:%M', ' %H:%M:%S']


def _make_date_object_(arg, fmt):
    if type(arg) == datetime:
        return arg
    else:
        if arg == 'today':
            return datetime.today()
        elif arg == 'now':
            return datetime.now()
        else:
            dfmt_str = strptime_fmt[dfmt]
            # now see if we have any time component
            tc = arg.count(':')
            dmft_str = dfmt_str + time_fmt[tc]
            return datetime.strptime(arg, dfmt_str)


def _datediff_(args):
    scale = {
        's': 1,
        'm': 60,
        'h': 3600,
        'd': 86400,
        'M': 26300160,
        'y': 960593618.9
    }

    # evaluate the branches.  Can't be sure how many we have
    arg_list = _list_args_(args)
    # now check the dateformates.
    if arg_list[3] in ['ymd', 'mdy', 'dmy']:
        dfmt = arg_list[3]  # got it
    else:
        dfmt = 'ymd'  # default

    date1 = _make_date_object_(args[0], dfmt)
    date2 = _make_date_object_(args[1], dfmt)
    datediff = date2 - date1  # time delta object
    days, seconds = datediff.days, datediff.seconds
    total_secs = days * 60 * 60 * 24 + seconds
    signed = False
    if type(arg_list[-1]) == bool:
        signed = arg_list[-1]
    if not signed:
        total_secs = abs(total_secs)

    units = arg_list[2]  # what units?
    result = total_secs / scale[units]
    return result

    return


def _sum_(args):
    return sum(_list_args_(args))


def _abs_(args):
    return abs(evaluateTree(args[0]))


def _min_(args):
    return min(_list_args_(args))


def _max_(args):
    return max(_list_args_(args))


def _sqrt_(args):
    return math.sqrt(evaluateTree(args[0]))


def _mean_(args):
    return statistics.mean(_list_args_(args))


def _median_(args):
    return statistics.median(_list_args_(args))


def _stdev_(args):
    return statistics.stdev(_list_args_(args))


def _roundup_(args):
    operand = evaluateTree(args[0])   # the number we're trying to round
    places = int(evaluateTree(args[1]))   # number of decimals
    a = operand * 10 ** places  # shift so we can use ceil functon
    return math.ceil(a) * 10 ** (- places)  # shift it back


def _rounddown_(args):
    operand = evaluateTree(args[0])   # the number we're trying to round
    places = int(evaluateTree(args[1]))   # number of decimals
    a = operand * 10 ** places  # shift so we can use ceil functon
    return math.floor(a) * 10 ** (- places)  # shift it back


def _round_(args):
    operand = evaluateTree(args[0])   # the number we're trying to round
    places = int(evaluateTree(args))   # number of decimals
    return round(operand, places)


def _if_(args):
    cond = evaluateTree(args[0])
    if cond:
        return evaluateTree(args[1])
    else:
        return evaluateTree(args[2])


def process_participant(args, data, dictionary, fem):
    """
    OUTERMOST LOOP: read through the dictionary one variable at a time
    MIDDLE LOOP: get the form name for the variable and loop through the
       form_event_map looking for entries that reference this form. There
       may be more than one if the form is mapped into more than one event.
    INNER LOOP: When we find the form name, loop through the returned data and
      process the value found in that event.  We will have to loop through all
       data because there may be multiple instances of the event
    """
    global Field_Value  # value of current variable
    global Current_Variable
    for Current_Variable, rec in dictionary.items():
        form = rec['form_name']

        # now we're going to loop through the form_event_table looking
        # at each event to see if it includes this form
        for form_event in fem:  # fem is a list of dictionaries

            # we've found a reference to this variable's form
            if form == form_event['form']:
                event = form_event['unique_event_name']

                # now go find that event in the REDCap data. Need to check
                # that the event name is right for this variable and also that
                # the form name matches the redcap repeat instrument. Latter
                # will only matter for repeating forms in non-repeating events.
                for entry in data:  # data is also a list of dictionaries
                    # matched the event.  now check for repeat instruments
                    if entry['redcap_event_name'] == event:
                        if form in is_repeat:
                            # list of events in which this repeats
                            elist = is_repeat[form]
                            if entry['redcap_event_name'] in elist:
                                if form != entry['redcap_repeat_instrument']:
                                    continue  # get out of here
                        elif entry['redcap_repeat_instrument'] != '':
                            continue

                        # field_value = process_fields(var, entry, dictionary)
                        Field_Value = return_redcap_var(Current_Variable,
                                                        entry, dictionary)
                        if Field_Value is None:
                            sys.exit('Parser failure')

                        branch_str = dictionary[Current_Variable]['branching_logic']
                        if branch_str != '':
                            tree = parseExpression(branch_str)
                            branch = evaluateTree(tree, entry, dictionary)
                        # if branch:  # these are the records we want
                        #     if rec['black_list']:

                        #         black_list = rec['black_list'].split('|')
                        #         if len(Field_Value) > 0:
                        #             check = Field_Value[0]
                        #         else:
                        #             check = ''

                        #         if check in black_list:
                        #             out_write.
                        #             writerow([entry['participationid'],
                        #                       entry['redcap_event_name'],
                        #                       entry['redcap_repeat_instance'],
                        #                       Current_Variable,
                        #                       'Missing Value', Field_Value])

    return


def return_data(args, big_data):
    """
    return a list of records for a single participant.  need to pass an
    iterator of big_data.  big_data doesn't
    need to be sorted in any particular order but records for a given
    participant must be adjacent
    """
    def _participant_(rec):
        """function to return the participantid for this record"""
        return rec['participationid']

    grouped_data = itertools.groupby(big_data, _participant_)
    for key, gen in grouped_data:

        data = list(gen)
        scans = 0
        selector = []  # will use to filter data later
        for rec in data:
            selector.append(True)
            # void participant
            if args.void and rec['redcap_event_name'] ==\
                    'administrative_inf_arm_1' and \
                    rec['void_participant'] == '1':
                print('participant {} is a void record'.
                      format(rec['participationid']))
                data = []  # chuck everything out
                break  # we're done with this participant

            # check the scan records for void or disabled
            if rec['redcap_event_name'] in ['neonatal_scan_arm_1',
                                            'fetal_scan_arm_1']:
                if args.disabled and rec['scan_disabled'] == '1':
                    selector[-1] = False  # de-select this record
                    print('participant {} {} {} marked disabled'.
                          format(rec['participationid'],
                                 rec['redcap_event_name'],
                                 rec['redcap_repeat_instance']))
                elif args.pilot == 0 and rec['scan_pilot'] == '1':
                    selector[-1] = False  # de-select this record
                    print('participant {} {} {} is a pilot scan'.
                          format(rec['participationid'],
                                 rec['redcap_event_name'],
                                 rec['redcap_repeat_instance']))
                else:
                    scans += 1  # found a good scan

        if len(data) <= 0:
            continue  # would have been a void and we've already printed
        elif args.noscan and scans == 0:
            print('participant {} has no usable scans'.
                  format(rec['participationid']))
            #  data = []  # clear
            continue
        else:
            data = list(itertools.compress(data, selector))
            yield data


def BuildDictionary(args, meta):
    """
    build a dictionary of all variables in which we're interested.  This will
    combine input file with the project metadata extracted from REDCap.  The
    output will contain all the project data but only for variables specified
    in the input file.  The infile will also contribute test logic expressions
    to be used in validating the field.  Will return as a dictionary of
    dictionaries.  Input file can be Excel file or .csv
    """
    infile = args.dictionary
    _fpat_ = re.compile(r'\s*(?P<file>.*xls[xmb]?)\s*(\[\s*(?P<sheet>.*)\b)')
    match = _fpat_.match(infile)  # test if filespec is Excel

    '''
    Build a dictionary of dictionaries.  Top level uses the variable name
    from 1st column as a key with value equal to a dictionary of all the
    other variables in the row with each getting a key equal to the column
    heading
    '''
    dictionary = {}
    if match:  # if we matched then it's an Excel file
        print('reading Excel File')
        # read the data dictionary
        dict_cols = {}
        infile = load_workbook(match.group('file'))
        if match.group('sheet'):
            source = infile[match.group('sheet')]
        else:
            source = infile.active
        dg = source.iter_rows(min_row=1, values_only=True)
        header = next(dg, None)  # get the header row

        # if we have a field called 'Ignore' then flag which column
        ig_col = header.index('Ignore') if 'Ignore' in header else -999
        # then loop through the data rows.  Will return a tuple each time
        for rec in dg:
            if ig_col >= 0:
                if rec[ig_col] in ['Yes', 1, '1', True]:
                    continue
            dic_entry = {}
            for i in range(1, len(header)):
                if rec[i] is None:
                    dic_entry[header[i].lower().replace(' ', '_')] = ''
                else:
                    dic_entry[header[i].lower().replace(' ', '_')] = rec[i]

            dictionary[rec[0]] = dic_entry

        infile.close()

    else:
        print('reading .csv file')
        with open(infile, 'r') as f:
            inreader = csv.reader(f)
            headers = next(inreader, None)
            ig_col = header.index('Ignore') if 'Ignore' in header else -999
            dictionary = {}
            for rec in inreader:
                if ig_col > 0:
                    if rec[ig_col] in ['Yes', 1, '1', True]:
                        continue
                dic_entry = {}
                for i in range(1, len(header)):
                    dic_entry[header[i].lower().replace(' ', '_')] = rec[i]
                dictionary[rec[0]] = dic_entry

    # now merge the project meta into dictionary
    for row in meta:
        if row['field_name'] in dictionary:
            for key, value in row.items():
                if (key != 'field_name' and
                        not (args.xlimits and
                             key.startswith('text_validation'))):
                    dictionary[row['field_name']][key] = value

    """
    Check date format.  We can distinguish betweem date and datetime but we
    don't have any good way of distinguishing between day-month-year and
    month-day-year. To address scan the dictionary looking for the text
    validation values.  We have to assume that we won't have a mixture of US
    and normal dates in the same project.
    """
    global American_Date
    American_Date = False
    for r in dictionary.values():
        if r['field_type'] == 'text':
            ft = r['text_validation_type_or_show_slider_number']
            if ft:
                if ft.find('mdy') >= 0:
                    American_Date = True

    return dictionary


# main program
global is_repeat  # used in process_participant
global _checkbox_  # will be used in return_redcap_var
_checkbox_ = re.compile(
    r'(\S+)\(([A-Za-z0-9-]+)\)|(\S+?)(\_{3,4}[A-Za-z0-9-]+$)')
is_repeat = {
    'dna_sample': ['baby_born_arm_1',
                   '18_month_assessment_arm_1'],
    'post_scan_events': ['post_scan_event_arm_1']
}

parser = argparse.ArgumentParser(description='RedCap REST client')
parser.add_argument('--dictionary', type=str,
                    default='Codebook for Error Check.xlsx [Working Copy]',
                    help='''Dictionary file with variables to check
                    If dictionary file is an Excel workbook then specification
                    can be followed by an optional worksheet name enclosed in
                    square brackets ([]).  If no worksheet name is given then
                    the first sheet in the workbook will be used.  If file
                    does not have an extension .xlx, .xlsx, .xlsb or .xlsm
                    then it will be processed as an ordinary comma seperated
                    test file''')
parser.add_argument('--output', type=argparse.FileType('w'),
                    default='dHCPmissing.txt',
                    help='Output file')
parser.add_argument('--pilot', action='count', default=0,
                    help='''--p means include pilot scans in deciding whether
                    to processs a record but ignore any data in it.
                    --pp means process the data from pilot scans too''')
parser.add_argument('--void', action='store_false',
                    help='include void participant records')
parser.add_argument('--reuse', action='store_true',
                    help='''don\'t reload REDCap DB if it\'s already present in
                    memory.  This will safe time during testing''')
parser.add_argument('--disabled', action='store_false',
                    help='include disabled scan records')
parser.add_argument('--noscan', action='store_false',
                    help='include records with no scans at all')
parser.add_argument('--xlimits', action='store_true',
                    help='''override maximum and minimum limits on variable
                    values in the meta data with values from the dictionary''')
parser.add_argument('records_of_interest', metavar='ID', type=str, nargs='*',
                    help='a list of subject IDs to fetch metadata from')
args = parser.parse_args()

# fetch API key from ~/.redcap-key ... don't keep in the source
key_filename = os.path.expanduser('~') + '/.redcap-key'
if not os.path.isfile(key_filename):
    print('redcap key file {} not found'.format(key_filename),
          file=sys.stderr)
    sys.exit(1)
api_key = open(key_filename, 'r').read().strip()

api_url = 'https://externalredcap.isd.kcl.ac.uk/api/'  # dhcp specific
project = Project(api_url, api_key)
if not os.path.isfile(key_filename):
    print('redcap key file {} not found'.format(key_filename))
    sys.exit(1)
api_key = open(key_filename, 'r').read().strip()

api_url = 'https://externalredcap.isd.kcl.ac.uk/api/'  # dHCP specific
project = Project(api_url, api_key)
meta = project.export_metadata()  # get the project metadata
dictionary = BuildDictionary(args, meta)
fields_of_interest = list(dictionary.keys())

"""
get the data from REDCap.  Will retrieve two lists:
    big_data is a list of all the data
    fem is the form event mapping
this takes some time so running interactively we can save time by not reading
again if the data already exists.  Will assume that if big_data exists then
so do meta and fem.  Note that if records_of_interest were specified on the
load and --reuse is specified then the record list doesn't get updated
"""

try:
    big_data  # does it exist?
except NameError:
    args.reuse = False  # force a load
if not args.reuse:
    print('loading REDCap data')

    fields_of_interest = list(dictionary.keys())
    try:
        big_data = project.export_records(fields=fields_of_interest,
                                          records=args.records_of_interest,
                                          format='json')
    except RedcapError:
        print('Redcap export too large', file=sys.stderr)
        sys.exit(1)

    fem = project.export_fem()                      # form event mappings

for data in return_data(args, big_data):
    if len(data) > 0:                     # have we got any?
        # yes: process them
        process_participant(args, data, dictionary, fem)

# out.close()  # close the output file
